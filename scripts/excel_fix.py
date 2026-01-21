
# ------ 
# Modify code for kode in string
# --------

import os
import pandas as pd
from openpyxl import load_workbook

def fix_entitas_nomor_aju_to_text(input_excel):
    if not os.path.isfile(input_excel):
        raise FileNotFoundError(f"Input file not found: {input_excel}")

    output_excel = input_excel  # overwrite same file

    wb = load_workbook(input_excel)
    sheet_names = wb.sheetnames

    sheets = {}
    for sheet in sheet_names:
        if sheet == "ENTITAS":
            # Existing logic for ENTITAS
            df = pd.read_excel(
                input_excel,
                sheet_name=sheet,
                dtype={
                    "NOMOR AJU": str,
                    "NOMOR IDENTITAS": str
                }
            )
        elif sheet == "HEADER":
            # FIX: Use converters to force Column C (index 2) and F (index 5) to str
            # This preserves '050900' as '050900'
            df = pd.read_excel(
                input_excel,
                sheet_name=sheet,
                converters={
                    2: str,  # Column C
                    5: str   # Column F
                }
            )
        else:
            df = pd.read_excel(input_excel, sheet_name=sheet)

        sheets[sheet] = df

    with pd.ExcelWriter(output_excel, engine="xlsxwriter") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet]
            text_format = workbook.add_format({"num_format": "@"})

            # Handle Formatting for ENTITAS
            if sheet == "ENTITAS":
                for col in ["NOMOR AJU", "NOMOR IDENTITAS"]:
                    if col in df.columns:
                        col_idx = df.columns.get_loc(col)
                        worksheet.set_column(col_idx, col_idx, None, text_format)

            # Handle Formatting for HEADER
            elif sheet == "HEADER":
                # Apply text format to Column C (2) and F (5) in the output file as well
                # Note: valid only if these columns exist in the dataframe
                for col_idx in [2, 5]: 
                    if col_idx < len(df.columns):
                        worksheet.set_column(col_idx, col_idx, None, text_format)

    return output_excel

# -------------
# # Standalone code runner
# # # --------------
# import os
# import pandas as pd
# from openpyxl import load_workbook

# def fix_excel_format(input_path, output_path):
#     if not os.path.isfile(input_path):
#         raise FileNotFoundError(f"Input file not found: {input_path}")

#     print(f"Processing file: {input_path}...")

#     # Load the workbook to get sheet names
#     wb = load_workbook(input_path, read_only=True)
#     sheet_names = wb.sheetnames
#     wb.close()

#     sheets = {}
    
#     # --- 1. READ DATA ---
#     for sheet in sheet_names:
#         if sheet == "ENTITAS":
#             # Logic for ENTITAS: Read specific columns as string
#             df = pd.read_excel(
#                 input_path,
#                 sheet_name=sheet,
#                 dtype={
#                     "NOMOR AJU": str,
#                     "NOMOR IDENTITAS": str
#                 }
#             )
#         elif sheet == "HEADER":
#             # Logic for HEADER: Force Column C (index 2) and F (index 5) to string
#             # This prevents '050900' from becoming 50900
#             df = pd.read_excel(
#                 input_path,
#                 sheet_name=sheet,
#                 converters={
#                     2: str,  # Column C
#                     5: str   # Column F
#                 }
#             )
#         else:
#             # Read other sheets normally
#             df = pd.read_excel(input_path, sheet_name=sheet)

#         sheets[sheet] = df

#     # --- 2. WRITE DATA ---
#     with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
#         for sheet, df in sheets.items():
#             df.to_excel(writer, sheet_name=sheet, index=False)

#             workbook = writer.book
#             worksheet = writer.sheets[sheet]
#             text_format = workbook.add_format({"num_format": "@"})

#             # Apply formatting for ENTITAS
#             if sheet == "ENTITAS":
#                 for col in ["NOMOR AJU", "NOMOR IDENTITAS"]:
#                     if col in df.columns:
#                         col_idx = df.columns.get_loc(col)
#                         worksheet.set_column(col_idx, col_idx, None, text_format)

#             # Apply formatting for HEADER
#             elif sheet == "HEADER":
#                 # Apply text format to Column C (index 2) and F (index 5)
#                 # We iterate through indices 2 and 5
#                 for col_idx in [2, 5]: 
#                     # Ensure column exists before applying format
#                     if col_idx < len(df.columns):
#                         worksheet.set_column(col_idx, col_idx, None, text_format)

#     print(f"Success! Output saved to: {output_path}")


# # --- MAIN EXECUTION BLOCK ---
# if __name__ == "__main__":
#     # 1. INPUT: Set the path to your input file here
#     # You can use r"" string to handle backslashes in Windows paths easily
#     input_file_path = r"/Users/anilpal/Documents/Inabata_Production/data/intermediate/00002701069420260118000303.xlsx" 

#     # 2. OUTPUT CONFIGURATION
#     # Get the directory where THIS script is currently located
#     script_dir = os.path.dirname(os.path.abspath(__file__))
    
#     # Create an output filename (e.g., "fixed_input_data.xlsx")
#     input_filename = os.path.basename(input_file_path)
#     output_filename = f"fixed_{input_filename}"
    
#     # Combine script directory with new filename to get full output path
#     output_file_path = os.path.join(script_dir, output_filename)

#     # 3. RUN THE PROCESS
#     try:
#         fix_excel_format(input_file_path, output_file_path)
#     except Exception as e:
#         print(f"An error occurred: {e}")