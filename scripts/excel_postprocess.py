import os
import pandas as pd
from openpyxl import load_workbook
from difflib import get_close_matches
from itertools import cycle
from datetime import datetime

# --------------------------------------------------
# HELPERS
# --------------------------------------------------
def get_col_indices(ws):
    return {
        str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    }


def format_date(value):
    if not value:
        return ""
    try:
        # Check if it's already a datetime object
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return value


# --------------------------------------------------
# CORE BUSINESS LOGIC
# --------------------------------------------------
# We updated the arguments here to match what run_pipeline.py sends
def process_customs_excel(input_excel_path, customer_ref_path, hs_code_path):
    
    # ---------- VALIDATION ----------
    print(f"   > Processing: {os.path.basename(input_excel_path)}")
    
    # Check if files exist
    for path, label in [
        (input_excel_path, "Generated Excel"),
        (customer_ref_path, "Customer Reference"),
        (hs_code_path, "HS Code Reference"),
    ]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"{label} file not found → {path}")

    # Load the Excel generated in Step 2
    wb = load_workbook(input_excel_path)

    # ---------- LOAD REFERENCES ----------
    # Load Customer List
    try:
        df_customers = pd.read_excel(customer_ref_path)
        df_customers.columns = df_customers.columns.astype(str).str.strip()
    except Exception as e:
        print(f"   ! Warning: Could not read Customer Ref: {e}")
        df_customers = pd.DataFrame()

    # Load HS Code List
    try:
        df_hs = pd.read_excel(hs_code_path)
        hs_map = dict(
            zip(
                df_hs["URAIAN"].astype(str).str.strip(),
                df_hs["HS"]
            )
        )
    except Exception as e:
        print(f"   ! Warning: Could not read HS Code Ref: {e}")
        hs_map = {}

    header_nomor_aju = None
    header_tanggal_pernyataan = None

    # ---------------------------------------------------------
    # SHEET: HEADER
    # ---------------------------------------------------------
    if "HEADER" in wb.sheetnames:
        ws = wb["HEADER"]
        cols = get_col_indices(ws)
        r = 2

        if "NOMOR AJU" in cols:
            header_nomor_aju = ws.cell(r, cols["NOMOR AJU"]).value

        if "TANGGAL PERNYATAAN" in cols:
            val = ws.cell(r, cols["TANGGAL PERNYATAAN"]).value
            header_tanggal_pernyataan = format_date(val)
            ws.cell(r, cols["TANGGAL PERNYATAAN"]).value = header_tanggal_pernyataan

        static_updates = {
            "KODE KANTOR": "050900",
            "KODE KANTOR TUJUAN": "050900",
            "KODE JENIS TPB": "2",
            "KODE TUJUAN PENGIRIMAN": "1",
            "KODE TUJUAN TPB": "1",
            "KOTA PERNYATAAN": "BEKASI",
            "NAMA PERNYATAAN": "NATAMIA",
            "JABATAN PERNYATAAN": "MANAGER",
        }

        for k, v in static_updates.items():
            if k in cols:
                ws.cell(r, cols[k]).value = v

    # ---------------------------------------------------------
    # SHEET: ENTITAS
    # ---------------------------------------------------------
    if "ENTITAS" in wb.sheetnames and not df_customers.empty:
        ws = wb["ENTITAS"]
        cols = get_col_indices(ws)

        ref_names = df_customers["NAMA ENTITAS"].dropna().astype(str).tolist()

        hardcoded_7 = {
            "NOMOR IDENTITAS": "0010694040059000000000",
            "NAMA ENTITAS": "INABATA INDONESIA",
            "ALAMAT ENTITAS": "KAWASAN INDUSTRI MM2100 JALAN BALI BLOK J-10, BEKASI",
        }

        hardcoded_3 = {
            "NOMOR IDENTITAS": "0010694040059000000000",
            "NAMA ENTITAS": "INABATA INDONESIA",
            "NIB ENTITAS": "9120101260717",
            "NOMOR IJIN ENTITAS": "3/KM.4/WBC.08/2025",
        }

        for r in range(2, ws.max_row + 1):
            kode = str(ws.cell(r, cols.get("KODE ENTITAS")).value or "").strip()
            
            # Logic for Customer (Kode 8)
            if kode == "8":
                name = ws.cell(r, cols.get("NAMA ENTITAS")).value
                if name:
                    # Fuzzy match the customer name
                    match = get_close_matches(str(name), ref_names, n=1, cutoff=0.6)
                    if match:
                        ref_row = df_customers[df_customers["NAMA ENTITAS"] == match[0]].iloc[0]
                        for col_name, col_idx in cols.items():
                            if col_name in ref_row.index:
                                ws.cell(r, col_idx).value = ref_row[col_name]

            # Logic for Sender (Kode 7)
            elif kode == "7":
                for k, v in hardcoded_7.items():
                    if k in cols:
                        ws.cell(r, cols[k]).value = v

            # Logic for Owner (Kode 3)
            elif kode == "3":
                for k, v in hardcoded_3.items():
                    if k in cols:
                        ws.cell(r, cols[k]).value = v

    # ---------------------------------------------------------
    # SHEET: DOKUMEN
    # ---------------------------------------------------------
    if "DOKUMEN" in wb.sheetnames:
        ws = wb["DOKUMEN"]
        cols = get_col_indices(ws)
        doc_cycle = cycle([380, 217, 630])

        for r in range(2, ws.max_row + 1):
            # Only process rows that look like they have data (check SERI)
            if "SERI" in cols and ws.cell(r, cols.get("SERI")).value:
                if header_nomor_aju and "NOMOR AJU" in cols:
                    ws.cell(r, cols["NOMOR AJU"]).value = header_nomor_aju
                if header_tanggal_pernyataan and "TANGGAL DOKUMEN" in cols:
                    ws.cell(r, cols["TANGGAL DOKUMEN"]).value = header_tanggal_pernyataan
                if "KODE DOKUMEN" in cols:
                    ws.cell(r, cols["KODE DOKUMEN"]).value = next(doc_cycle)
    # -------------------------------------------------------------
    # SHEET: PENGANGKUT (NEW)
    # ---------------------------------------------------------
    if "PENGANGKUT" in wb.sheetnames:
        ws = wb["PENGANGKUT"]
        cols = get_col_indices(ws)
        counter = 1
        
        # Ensure the main column exists to avoid errors
        if "NAMA PENGANGKUT" in cols:
            for row in range(2, ws.max_row + 1):
                nama_pengangkut = ws.cell(row=row, column=cols["NAMA PENGANGKUT"]).value
                
                if nama_pengangkut:
                    if "SERI" in cols:
                        ws.cell(row=row, column=cols["SERI"]).value = counter
                        counter += 1
                    if "NOMOR AJU" in cols and header_nomor_aju:
                        ws.cell(row=row, column=cols["NOMOR AJU"]).value = header_nomor_aju
                    if "NOMOR PENGANGKUT" in cols:
                        ws.cell(row=row, column=cols["NOMOR PENGANGKUT"]).value = "-"
        print("   > Processed PENGANGKUT sheet.")
    # ---------------------------------------------------------
    # SHEET: BARANG
    # ---------------------------------------------------------
    if "BARANG" in wb.sheetnames:
        ws = wb["BARANG"]
        cols = get_col_indices(ws)
        counter = 1

        for r in range(2, ws.max_row + 1):
            uraian = ws.cell(r, cols.get("URAIAN")).value
            if uraian:
                if "SERI BARANG" in cols:
                    ws.cell(r, cols["SERI BARANG"]).value = counter
                    counter += 1

                if header_nomor_aju and "NOMOR AJU" in cols:
                    ws.cell(r, cols["NOMOR AJU"]).value = header_nomor_aju

                # Lookup HS Code if missing
                if "HS" in cols: 
                    current_hs = ws.cell(r, cols["HS"]).value
                    if not current_hs:
                        key = str(uraian).strip()
                        if key in hs_map:
                            ws.cell(r, cols["HS"]).value = hs_map[key]

    # ---------------------------------------------------------
    # SAVE OUTPUT (Overwrite the intermediate file)
    # ---------------------------------------------------------
    # We overwrite the input file so the "final" file contains the updates
    wb.save(input_excel_path)
    return input_excel_path


# --------------------------------------------------
# ENTRY POINT
# --------------------------------------------------
if __name__ == "__main__":
    print("This script is intended to be run via scripts.run_pipeline")