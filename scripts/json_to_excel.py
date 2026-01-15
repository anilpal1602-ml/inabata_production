import json
import os
import yaml
from openpyxl import load_workbook
from datetime import datetime

# --------------------------------------------------
# LOAD CONFIG
# --------------------------------------------------
CONFIG_PATH = os.getenv("CONFIG_PATH", "config.yaml")

# Check if config exists before loading to prevent crash if run from wrong dir
if os.path.exists(CONFIG_PATH):
    with open(CONFIG_PATH, "r") as f:
        config = yaml.safe_load(f)
        BASE_DIR = config["base_dir"]
        OUTPUT_DIR = os.path.join(BASE_DIR, config["data"]["output"]["final_excel_dir"])
        SERIAL_TRACKER_PATH = os.path.join(BASE_DIR, "data/state/serial_tracker.txt")
else:
    # Fallback/Default if config not found immediately (helpful for testing)
    BASE_DIR = os.path.dirname(__file__)
    OUTPUT_DIR = os.path.join(BASE_DIR, "output")
    SERIAL_TRACKER_PATH = os.path.join(BASE_DIR, "serial_tracker.txt")

# --------------------------------------------------
# SERIAL NUMBER HANDLER
# --------------------------------------------------
def get_next_serial_number(tracker_path, fallback_serial):
    current_serial = 0

    if os.path.exists(tracker_path):
        try:
            with open(tracker_path, "r") as f:
                val = f.read().strip()
                if val.isdigit():
                    current_serial = int(val)
        except Exception:
            pass

    if current_serial == 0 and str(fallback_serial).isdigit():
        current_serial = int(fallback_serial)

    new_serial = current_serial + 1

    os.makedirs(os.path.dirname(tracker_path), exist_ok=True)
    with open(tracker_path, "w") as f:
        f.write(str(new_serial))

    return new_serial


# --------------------------------------------------
# CORE BUSINESS LOGIC
# --------------------------------------------------
def json_to_excel(json_path, template_path):
    # ---------- Validate Inputs ----------
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"JSON not found → {json_path}")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found → {template_path}")

    # ---------- Load JSON ----------
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # ---------- Load Excel Template ----------
    wb = load_workbook(template_path)
    generated_nomor_aju = None

    # ---------- NOMOR AJU LOGIC ----------
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        nomor_col = None

        # This was the line causing your error previously
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=col).value).strip() == "NOMOR AJU":
                nomor_col = col
                break

        if nomor_col:
            existing_val = str(ws.cell(row=2, column=nomor_col).value or "").strip()
            if len(existing_val) < 26:
                existing_val = "00002701069420200101000000"

            prefix = existing_val[:12]
            old_serial = existing_val[-6:]
            new_date = datetime.now().strftime("%Y%m%d")

            new_serial = get_next_serial_number(
                SERIAL_TRACKER_PATH,
                old_serial
            )

            generated_nomor_aju = f"{prefix}{new_date}{new_serial:06d}"
            ws.cell(row=2, column=nomor_col).value = generated_nomor_aju
            break

    # ---------- JSON → EXCEL ----------
    for sheet_name, content in data.items():
        if sheet_name not in wb.sheetnames or len(content) < 2:
            continue

        ws = wb[sheet_name]
        headers = content[0]
        rows = content[1:]

        excel_headers = {
            str(ws.cell(row=1, column=c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value
        }

        # ENTITAS SPECIAL CASE
        if sheet_name == "ENTITAS":
            kode_col = excel_headers.get("KODE ENTITAS")
            target_row = None

            if kode_col:
                for r in range(2, ws.max_row + 1):
                    if str(ws.cell(row=r, column=kode_col).value).strip() == "8":
                        target_row = r
                        break

            if target_row:
                for row in rows[:1]:
                    for i, val in enumerate(row):
                        col = excel_headers.get(headers[i])
                        if col:
                            ws.cell(row=target_row, column=col).value = val
            continue

        # STANDARD SHEETS
        excel_row = 2
        for row in rows:
            for i, val in enumerate(row):
                col = excel_headers.get(headers[i])
                if col:
                    ws.cell(row=excel_row, column=col).value = val
            excel_row += 1

    # ---------- OUTPUT ----------
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    filename = (
        f"{generated_nomor_aju}.xlsx"
        if generated_nomor_aju
        else f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    output_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(output_path)

    return output_path


# --------------------------------------------------
# PIPELINE ENTRY (USED BY run_pipeline.py)
# --------------------------------------------------
def run(json_path, template_path):
    output_file = json_to_excel(json_path, template_path)

    print("Excel generation successful")
    print(f"Output file → {output_file}")

    return output_file


# --------------------------------------------------
# ENTRY POINT
# --------------------------------------------------
if __name__ == "__main__":
    # Simple test execution if run directly
    print("This script is designed to be run via the pipeline.")